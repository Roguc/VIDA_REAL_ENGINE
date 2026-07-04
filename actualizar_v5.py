from pathlib import Path

ROOT = Path(__file__).resolve().parent

ARCHIVOS = {
    "Core/logger.py": '''
from datetime import datetime

def log(texto):
    ahora = datetime.now().strftime("%H:%M:%S")
    print(f"[{ahora}] {texto}")
''',

    "Models/sistema_dia.py": '''
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Dict, List

@dataclass
class SistemaDia:
    fecha: date | None = None
    fecha_texto: str = ""
    hora_actual: str = ""
    tipo_dia: str = ""
    turno_serpat: str = ""

    identidad: Dict[str, Any] = field(default_factory=dict)
    contexto: Dict[str, Any] = field(default_factory=dict)

    universidad: Dict[str, Any] = field(default_factory=dict)
    salud: Dict[str, Any] = field(default_factory=dict)
    finanzas: Dict[str, Any] = field(default_factory=dict)
    empresas: Dict[str, Any] = field(default_factory=dict)
    serpat: Dict[str, Any] = field(default_factory=dict)
    ancla: Dict[str, Any] = field(default_factory=dict)
    continuidad: Dict[str, Any] = field(default_factory=dict)
    recursos: Dict[str, Any] = field(default_factory=dict)

    pendientes: List[Dict[str, Any]] = field(default_factory=list)
    alertas: List[Dict[str, Any]] = field(default_factory=list)
    decisiones: List[Dict[str, Any]] = field(default_factory=list)
    cronograma: List[Dict[str, Any]] = field(default_factory=list)

    def agregar_alerta(self, area, prioridad, titulo, detalle):
        self.alertas.append({
            "area": area,
            "prioridad": prioridad,
            "titulo": titulo,
            "detalle": detalle
        })

    def agregar_decision(self, area, prioridad, accion, motivo, registro=""):
        self.decisiones.append({
            "area": area,
            "prioridad": prioridad,
            "accion": accion,
            "motivo": motivo,
            "registro": registro
        })

    def ordenar_decisiones(self):
        self.decisiones.sort(key=lambda x: x.get("prioridad", 0), reverse=True)
''',

    "Managers/excel_manager.py": '''
from pathlib import Path
from openpyxl import load_workbook

class ExcelManager:
    def __init__(self):
        self.archivo = None
        self.workbook = None
        self.hojas = []

    def buscar_excel(self):
        carpeta = Path("Excel")
        archivos = list(carpeta.glob("*.xlsx"))
        if not archivos:
            raise FileNotFoundError("No se encontró ningún Excel en la carpeta Excel.")
        archivos.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        self.archivo = archivos[0]
        return self.archivo

    def cargar(self):
        self.buscar_excel()
        self.workbook = load_workbook(self.archivo, data_only=True)
        self.hojas = self.workbook.sheetnames
        return True

    def existe_hoja(self, nombre):
        return nombre in self.hojas

    def obtener_hoja(self, nombre):
        if self.existe_hoja(nombre):
            return self.workbook[nombre]
        return None

    def leer_celda(self, hoja, celda):
        ws = self.obtener_hoja(hoja)
        if ws is None:
            return None
        return ws[celda].value

    def resumen(self):
        return {
            "archivo": self.archivo.name if self.archivo else "",
            "total_hojas": len(self.hojas),
            "hojas": self.hojas
        }
''',

    "Managers/recursos_manager.py": '''
from pathlib import Path

class RecursosManager:
    def __init__(self):
        self.carpeta = Path("Recursos")
        self.indice = {}

    def cargar(self):
        self.indice = {}
        if not self.carpeta.exists():
            return self.indice

        for carpeta in sorted(self.carpeta.iterdir()):
            if carpeta.is_dir():
                archivos = [a for a in carpeta.rglob("*") if a.is_file()]
                self.indice[carpeta.name] = {
                    "ruta": str(carpeta),
                    "total_archivos": len(archivos),
                    "archivos": [str(a) for a in archivos]
                }

        return self.indice

    def resumen(self):
        total = sum(v["total_archivos"] for v in self.indice.values())
        return {
            "total_categorias": len(self.indice),
            "total_archivos": total,
            "categorias": self.indice
        }
''',

    "Core/motor_central.py": '''
from datetime import datetime

from Models.sistema_dia import SistemaDia
from Managers.excel_manager import ExcelManager
from Managers.recursos_manager import RecursosManager


class MotorCentral:
    def __init__(self):
        self.sistema = SistemaDia()
        self.excel = ExcelManager()
        self.recursos = RecursosManager()

    def ejecutar(self):
        hoy = datetime.now()

        self.sistema.fecha = hoy.date()
        self.sistema.fecha_texto = hoy.strftime("%A %d-%m-%Y")
        self.sistema.hora_actual = hoy.strftime("%H:%M")

        self.excel.cargar()
        self.recursos.cargar()

        self.sistema.contexto["excel"] = self.excel.resumen()
        self.sistema.recursos = self.recursos.resumen()

        self._analizar_contexto_basico()
        self._generar_decisiones_base()

        self.sistema.ordenar_decisiones()

        return self.sistema

    def _analizar_contexto_basico(self):
        hojas = self.sistema.contexto["excel"]["hojas"]

        if "SERPAT TURNOS" in hojas:
            self.sistema.agregar_alerta(
                "SERPAT",
                80,
                "Calendario laboral detectado",
                "El sistema puede utilizar SERPAT TURNOS para adaptar el día."
            )

        if "PENDIENTES" in hojas:
            self.sistema.agregar_alerta(
                "Pendientes",
                90,
                "Hoja de pendientes detectada",
                "Debe integrarse al Motor Central en la siguiente versión."
            )

        if "VISION_BOARD" in hojas:
            self.sistema.agregar_alerta(
                "Identidad",
                70,
                "Vision Board detectado",
                "Debe usarse para la visualización diaria."
            )

    def _generar_decisiones_base(self):
        self.sistema.agregar_decision(
            "Sistema",
            100,
            "Abrir Centro de Comando y revisar prioridades del día.",
            "El Motor Central ya está activo.",
            "20_Registro_Diario; 21_KPI_Diario"
        )

        self.sistema.agregar_decision(
            "Universidad",
            95,
            "Revisar evaluaciones críticas y contenido pendiente.",
            "La universidad requiere pasar de bloque genérico a instrucción específica.",
            "61_T2_Ramos; 62_T2_Evaluaciones; 64_T2_Errores"
        )

        self.sistema.agregar_decision(
            "Salud",
            85,
            "Registrar presión, energía, sueño, medicamento e hidratación.",
            "El sistema debe proteger capital biológico.",
            "H02_Registro_Salud; H03_Presión_Log"
        )

        self.sistema.agregar_decision(
            "Finanzas",
            80,
            "Registrar movimientos, revisar gastos y actualizar caja.",
            "El sistema debe mantener visibilidad financiera.",
            "11_Ingresos; 12_Gastos; 13_Deudas; 18_Conciliacion_Bancaria_V3"
        )

        self.sistema.agregar_decision(
            "Empresas",
            75,
            "Ejecutar una acción empresarial real con evidencia.",
            "MGC, LNH o CaptaPropIA no deben quedar sin avance.",
            "30_MGC_CRM; 33_Seguimientos; 76_CaptaPropIA"
        )
''',

    "main.py": '''
from Core.motor_central import MotorCentral

def main():
    motor = MotorCentral()
    sistema = motor.ejecutar()

    print()
    print("===================================")
    print(" VIDA REAL ENGINE V5.1")
    print(" MOTOR CENTRAL IA — ACTIVO")
    print("===================================")
    print()
    print("Fecha:", sistema.fecha_texto)
    print("Hora:", sistema.hora_actual)
    print("Excel:", sistema.contexto["excel"]["archivo"])
    print("Hojas:", sistema.contexto["excel"]["total_hojas"])
    print("Recursos:", sistema.recursos["total_archivos"], "archivos")
    print()
    print("DECISIONES DEL MOTOR CENTRAL:")
    print()

    for i, decision in enumerate(sistema.decisiones, start=1):
        print(f"{i}. [{decision['area']}] Prioridad {decision['prioridad']}")
        print("   Acción:", decision["accion"])
        print("   Motivo:", decision["motivo"])
        print("   Registro:", decision["registro"])
        print()

    print("ALERTAS:")
    print()

    for alerta in sistema.alertas:
        print(f"- [{alerta['area']}] {alerta['titulo']}")
        print(" ", alerta["detalle"])
        print()

if __name__ == "__main__":
    main()
'''
}

def escribir_archivo(ruta_relativa, contenido):
    ruta = ROOT / ruta_relativa
    ruta.parent.mkdir(parents=True, exist_ok=True)
    ruta.write_text(contenido.strip() + "\\n", encoding="utf-8")
    print(f"OK: {ruta_relativa}")

def main():
    print()
    print("===================================")
    print(" ACTUALIZADOR VIDA REAL ENGINE V5.1")
    print("===================================")
    print()

    for ruta, contenido in ARCHIVOS.items():
        escribir_archivo(ruta, contenido)

    print()
    print("Actualización V5.1 instalada correctamente.")
    print("Ahora ejecuta: python main.py")
    print()

if __name__ == "__main__":
    main()