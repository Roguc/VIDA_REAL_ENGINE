from pathlib import Path
from textwrap import dedent

ROOT = Path(__file__).resolve().parent

FILES = {
    "Core/logger.py": r"""
from datetime import datetime

def log(texto: str) -> None:
    ahora = datetime.now().strftime("%H:%M:%S")
    print(f"[{ahora}] {texto}")
""",

    "Models/sistema_dia.py": r"""
from dataclasses import dataclass, field
from datetime import date
from typing import Any, Dict, List

@dataclass
class SistemaDia:
    fecha: date | None = None
    fecha_texto: str = ""
    hora_actual: str = ""
    tipo_dia: str = ""
    turno_serpat: str = ""

    identidad: Dict[str, Any] = field(default_factory=dict)
    contexto: Dict[str, Any] = field(default_factory=dict)

    universidad: Dict[str, Any] = field(default_factory=dict)
    salud: Dict[str, Any] = field(default_factory=dict)
    finanzas: Dict[str, Any] = field(default_factory=dict)
    empresas: Dict[str, Any] = field(default_factory=dict)
    serpat: Dict[str, Any] = field(default_factory=dict)
    ancla: Dict[str, Any] = field(default_factory=dict)
    continuidad: Dict[str, Any] = field(default_factory=dict)
    recursos: Dict[str, Any] = field(default_factory=dict)

    pendientes: List[Dict[str, Any]] = field(default_factory=list)
    alertas: List[Dict[str, Any]] = field(default_factory=list)
    decisiones: List[Dict[str, Any]] = field(default_factory=list)
    cronograma: List[Dict[str, Any]] = field(default_factory=list)

    def agregar_alerta(self, area: str, prioridad: int, titulo: str, detalle: str) -> None:
        self.alertas.append({
            "area": area,
            "prioridad": prioridad,
            "titulo": titulo,
            "detalle": detalle,
        })

    def agregar_decision(self, area: str, prioridad: int, accion: str, motivo: str, registro: str = "") -> None:
        self.decisiones.append({
            "area": area,
            "prioridad": prioridad,
            "accion": accion,
            "motivo": motivo,
            "registro": registro,
        })

    def ordenar_decisiones(self) -> None:
        self.decisiones.sort(key=lambda x: x.get("prioridad", 0), reverse=True)
""",

    "Managers/excel_manager.py": r"""
from pathlib import Path
from openpyxl import load_workbook

class ExcelManager:
    def __init__(self):
        self.archivo = None
        self.workbook = None
        self.hojas = []

    def buscar_excel(self):
        carpeta = Path("Excel")
        archivos = list(carpeta.glob("*.xlsx"))
        if not archivos:
            raise FileNotFoundError("No se encontró ningún Excel en la carpeta Excel.")
        archivos.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        self.archivo = archivos[0]
        return self.archivo

    def cargar(self):
        self.buscar_excel()
        self.workbook = load_workbook(self.archivo, data_only=True)
        self.hojas = self.workbook.sheetnames
        return True

    def existe_hoja(self, nombre: str) -> bool:
        return nombre in self.hojas

    def obtener_hoja(self, nombre: str):
        if self.existe_hoja(nombre):
            return self.workbook[nombre]
        return None

    def leer_celda(self, hoja: str, celda: str):
        ws = self.obtener_hoja(hoja)
        if ws is None:
            return None
        return ws[celda].value

    def resumen(self):
        return {
            "archivo": self.archivo.name if self.archivo else "",
            "total_hojas": len(self.hojas),
            "hojas": self.hojas,
        }
""",

    "Managers/recursos_manager.py": r"""
from pathlib import Path

class RecursosManager:
    def __init__(self):
        self.carpeta = Path("Recursos")
        self.indice = {}

    def cargar(self):
        self.indice = {}
        if not self.carpeta.exists():
            return self.indice

        for carpeta in sorted(self.carpeta.iterdir()):
            if carpeta.is_dir():
                archivos = [a for a in carpeta.rglob("*") if a.is_file()]
                self.indice[carpeta.name] = {
                    "ruta": str(carpeta),
                    "total_archivos": len(archivos),
                    "archivos": [str(a) for a in archivos],
                }

        return self.indice

    def resumen(self):
        total = sum(v["total_archivos"] for v in self.indice.values())
        return {
            "total_categorias": len(self.indice),
            "total_archivos": total,
            "categorias": self.indice,
        }
""",

    "Core/motor_central.py": r"""
from datetime import datetime

from Models.sistema_dia import SistemaDia
from Managers.excel_manager import ExcelManager
from Managers.recursos_manager import RecursosManager

class MotorCentral:
    def __init__(self):
        self.sistema = SistemaDia()
        self.excel = ExcelManager()
        self.recursos = RecursosManager()

    def ejecutar(self):
        ahora = datetime.now()

        self.sistema.fecha = ahora.date()
        self.sistema.fecha_texto = ahora.strftime("%d-%m-%Y")
        self.sistema.hora_actual = ahora.strftime("%H:%M")

        self.excel.cargar()
        self.recursos.cargar()

        self.sistema.contexto["excel"] = self.excel.resumen()
        self.sistema.recursos = self.recursos.resumen()

        self._analizar_contexto()
        self._generar_decisiones_base()
        self.sistema.ordenar_decisiones()

        return self.sistema

    def _analizar_contexto(self):
        hojas = self.sistema.contexto["excel"]["hojas"]

        if "SERPAT TURNOS" in hojas:
            self.sistema.agregar_alerta(
                "SERPAT",
                90,
                "Calendario laboral detectado",
                "Usar SERPAT TURNOS para adaptar el día."
            )

        if "PENDIENTES" in hojas:
            self.sistema.agregar_alerta(
                "Pendientes",
                95,
                "Pendientes detectados",
                "Integrar pendientes en la priorización diaria."
            )

        if "VISION_BOARD" in hojas:
            self.sistema.agregar_alerta(
                "Identidad",
                70,
                "Vision Board detectado",
                "Usar Vision Board para visualización diaria."
            )

        if "DESARROLLO_PERSONAL" in hojas:
            self.sistema.agregar_alerta(
                "Desarrollo Personal",
                75,
                "Desarrollo personal detectado",
                "Integrar lectura, hábitos e identidad."
            )

    def _generar_decisiones_base(self):
        self.sistema.agregar_decision(
            "Sistema",
            100,
            "Abrir Centro de Comando y revisar prioridades.",
            "Motor Central IA activo.",
            "20_Registro_Diario; 21_KPI_Diario"
        )
        self.sistema.agregar_decision(
            "Universidad",
            95,
            "Revisar evaluación crítica, contenidos, errores y próxima entrega.",
            "Universidad debe pasar de bloque genérico a instrucción específica.",
            "61_T2_Ramos; 62_T2_Evaluaciones; 64_T2_Errores"
        )
        self.sistema.agregar_decision(
            "Salud",
            85,
            "Registrar presión, energía, sueño, medicamento e hidratación.",
            "Proteger capital biológico.",
            "H02_Registro_Salud; H03_Presión_Log"
        )
        self.sistema.agregar_decision(
            "Finanzas",
            80,
            "Registrar movimientos, revisar gastos y actualizar caja.",
            "Mantener visibilidad financiera.",
            "11_Ingresos; 12_Gastos; 13_Deudas; 18_Conciliacion_Bancaria_V3"
        )
        self.sistema.agregar_decision(
            "Empresas",
            75,
            "Ejecutar una acción empresarial real con evidencia.",
            "MGC, LNH o CaptaPropIA no deben quedar sin avance.",
            "30_MGC_CRM; 33_Seguimientos; 76_CaptaPropIA"
        )
""",

    "main.py": r"""
from Core.motor_central import MotorCentral

def main():
    motor = MotorCentral()
    sistema = motor.ejecutar()

    print()
    print("===================================")
    print(" VIDA REAL ENGINE V5.1")
    print(" MOTOR CENTRAL IA — ACTIVO")
    print("===================================")
    print()
    print("Fecha:", sistema.fecha_texto)
    print("Hora:", sistema.hora_actual)
    print("Excel:", sistema.contexto["excel"]["archivo"])
    print("Hojas:", sistema.contexto["excel"]["total_hojas"])
    print("Recursos:", sistema.recursos["total_archivos"], "archivos")
    print()
    print("DECISIONES DEL MOTOR CENTRAL:")
    print()

    for i, decision in enumerate(sistema.decisiones, start=1):
        print(f"{i}. [{decision['area']}] Prioridad {decision['prioridad']}")
        print("   Acción:", decision["accion"])
        print("   Motivo:", decision["motivo"])
        print("   Registro:", decision["registro"])
        print()

    print("ALERTAS:")
    print()

    for alerta in sistema.alertas:
        print(f"- [{alerta['area']}] {alerta['titulo']}")
        print(" ", alerta["detalle"])
        print()

if __name__ == "__main__":
    main()
""",
}

def write_file(relative_path: str, content: str):
    path = ROOT / relative_path
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(dedent(content).strip() + "\n", encoding="utf-8")
    print(f"OK: {relative_path}")

def main():
    print()
    print("===================================")
    print(" ACTUALIZADOR VIDA REAL ENGINE V5.1")
    print("===================================")
    print()

    for relative_path, content in FILES.items():
        write_file(relative_path, content)

    print()
    print("Actualización V5.1 instalada.")
    print("Ejecuta ahora: python main.py")
    print()

if __name__ == "__main__":
    main()
