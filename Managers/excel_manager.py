from __future__ import annotations

import re
import unicodedata
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook


class ExcelManager:
    ALIASES_CRITICOS = {
        "SERPAT TURNOS": ["SERPAT TURNOS", "SERPAT_TURNOS", "TURNOS SERPAT"],
        "20_Registro_Diario": ["20_REGISTRO_DIARIO", "REGISTRO DIARIO"],
        "21_KPI_Diario": ["21_KPI_DIARIO", "KPI DIARIO"],
        "23_Domingo_CEO": ["23_DOMINGO_CEO", "DOMINGO CEO"],
        "44_Salud": ["44_SALUD", "SALUD"],
        "61_T2_Ramos": ["61_T2_RAMOS", "T2 RAMOS"],
        "62_T2_Evaluaciones": ["62_T2_EVALUACIONES", "T2 EVALUACIONES", "EVALUACIONES"],
        "63_T2_Bloques": ["63_T2_BLOQUES", "T2 BLOQUES"],
        "64_T2_Errores": ["64_T2_ERRORES", "T2 ERRORES", "ERRORES"],
        "DESARROLLO_PERSONAL": ["DESARROLLO_PERSONAL", "DESARROLLO PERSONAL"],
        "VISION_BOARD": ["VISION_BOARD", "VISION BOARD"],
        "PENDIENTES": ["PENDIENTES"],
        "30_MGC_CRM": ["30_MGC_CRM", "MGC CRM"],
        "33_Seguimientos": ["33_SEGUIMIENTOS", "SEGUIMIENTOS"],
        "70_Mercado_MGC": ["70_MERCADO_MGC", "MERCADO MGC"],
        "71_Mercado_LNR": ["71_MERCADO_LNR", "MERCADO LNR"],
        "76_CaptaPropIA": ["76_CAPTAPROPIA", "CAPTAPROPIA"],
        "82_LNH_Corporate_System": ["82_LNH_CORPORATE_SYSTEM", "LNH CORPORATE SYSTEM"],
    }

    CATEGORIAS = {
        "serpat": ["SERPAT TURNOS"],
        "registro": ["20_Registro_Diario", "21_KPI_Diario", "23_Domingo_CEO"],
        "finanzas": ["20_Registro_Diario", "21_KPI_Diario"],
        "salud": ["44_Salud"],
        "universidad": ["61_T2_Ramos", "62_T2_Evaluaciones", "63_T2_Bloques", "64_T2_Errores"],
        "desarrollo": ["DESARROLLO_PERSONAL", "VISION_BOARD"],
        "pendientes": ["PENDIENTES"],
        "empresas": [
            "30_MGC_CRM",
            "33_Seguimientos",
            "70_Mercado_MGC",
            "71_Mercado_LNR",
            "76_CaptaPropIA",
            "82_LNH_Corporate_System",
        ],
    }

    KEYWORDS_HEADERS = {
        "id",
        "area",
        "subarea",
        "tarea",
        "prioridad",
        "estado",
        "fecha",
        "turno",
        "ingreso",
        "salida",
        "categoria",
        "objetivo",
        "actividad",
        "ramo",
        "evaluacion",
        "observaciones",
        "avance",
        "registro",
    }

    DOMINIOS_FUNCIONALES = [
        "Registro Diario",
        "Finanzas",
        "Inversiones",
        "Universidad",
        "Salud",
        "Empresas",
        "SERPAT",
        "Desarrollo Personal",
        "Vision Board",
        "Ancla",
        "Alertas",
        "Pendientes",
        "KPI",
        "Recursos",
        "Historial",
        "Otros",
    ]

    DOMINIO_KEYWORDS = {
        "Registro Diario": ["REGISTRO", "DIARIO", "DOMINGO_CEO", "APERTURA"],
        "Finanzas": ["INGRESOS", "GASTOS", "DEUDAS", "CONCILIACION", "FINANZ", "CAJA", "BANCO", "PRESUPUEST"],
        "Inversiones": ["INVERSION", "ETF", "APV", "TQQQ", "SOXX", "BROKER", "PORTAFOLIO", "CARTERA", "RENTABILIDAD", "PATRIMONIO"],
        "Universidad": ["UNIVERSIDAD", "RAMO", "EVALUACION", "EVALUACIONES", "BLOQUES", "ERRORES", "ACADEM", "CALCULO", "FISICA", "ESTADIST"],
        "Salud": ["SALUD", "PRESION", "MEDICAMENTO", "SUENO", "HIDRATACION", "PULSO", "ENERGIA"],
        "Empresas": ["MGC", "LNR", "LNH", "CRM", "SEGUIMIENTOS", "CAPTAPROPIA", "MERCADO", "EMPRESA", "VENTA", "NEGOCIO"],
        "SERPAT": ["SERPAT", "TURNO", "TURNOS", "NOCTUR", "COLACION", "TRABAJADAS"],
        "Desarrollo Personal": ["DESARROLLO_PERSONAL", "MISION", "VISION", "PROPOSITO", "LEY_001", "HABITOS", "LECTURA"],
        "Vision Board": ["VISION_BOARD", "OBJETIVO_2042", "CATEGORIA", "CARTEL", "ATRA"],
        "Ancla": ["ANCLA", "IDENTIDAD", "CONSTRUCTOR", "SOBREVIVIENTE"],
        "Alertas": ["ALERTA", "CORRECCION", "RIESGO", "AUDITORIA"],
        "Pendientes": ["PENDIENTE", "PENDIENTES", "TAREA", "SUBAREA"],
        "KPI": ["KPI", "CUMPLIMIENTO", "SCORE", "INDICADOR"],
        "Recursos": ["RECURSO", "INDICE", "MAPA_OPERATIVO", "FUENTE"],
        "Historial": ["HISTORIAL", "RESPALDO", "LOG", "EVOLUCION", "TRACKING"],
    }

    EXCEL_DOMAIN_OVERRIDES = {
        "11_INGRESOS": "Finanzas",
        "40_ESTUDIO_EMPRESAS": "Empresas",
        "41_CLASIFICACION_EMPRESAS": "Empresas",
        "42_MAPA_MERCADO": "Empresas",
        "43_GUIA_OPORTUN": "Empresas",
        "77_AUDITORIA_LNR_MERCADO": "Empresas",
        "39_FLUJO_EMPRESAS_V6_4": "Empresas",
        "H00_DASHBOARD_HEALTH": "Salud",
        "H02_REGISTRO_SALUD": "Salud",
        "H03_DOCUMENTOS": "Salud",
        "H04_SEGUROS_BENEFICIOS": "Salud",
        "51_SEGUROS_Y_BENEFICIOS": "Salud",
        "52_DASHBOARD_HEALTH": "Salud",
        "53_HISTORIAL_DENTAL": "Salud",
        "54_HISTORIAL_FAMILIAR": "Salud",
    }

    PREFIX_DOMAIN_RULES = [
        (r"^H\d{2}_", "Salud"),
        (r"^(06)_", "SERPAT"),
        (r"^(20|23)_", "Registro Diario"),
        (r"^(21|22|24)_", "KPI"),
        (r"^(11|12|13|18|19)_", "Finanzas"),
        (r"^(07|08|09|10|14|15|16|17)_", "Inversiones"),
        (r"^(30|31|32|33|34|35|36|37|38|39|40|41|42|43|70|71|72|73|74|75|76|77|78|79|80|81|82|83)_", "Empresas"),
        (r"^(44|45|46|47|48|49|50|51|52|53|54)_", "Salud"),
        (r"^(55|56|57|58|59|60|61|62|63|64|65|66|67|68|69)_", "Universidad"),
        (r"^(90|91|92|93|94|95|96|97|98|99)_", "Alertas"),
    ]

    def __init__(self):
        self.archivo = None
        self.workbook = None
        self.hojas = []
        self.scan = {}
        self.alias_hojas = {}
        self.hojas_vacias = []
        self.hojas_problema = []
        self.catalogo = []
        self.dominios = {k: [] for k in self.DOMINIOS_FUNCIONALES}
        self.advertencias = []

    def buscar_excel(self):
        carpeta = Path("Excel")
        archivos = list(carpeta.glob("*.xlsx"))
        if not archivos:
            raise FileNotFoundError("No se encontro ningun Excel en la carpeta Excel.")
        archivos.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        self.archivo = archivos[0]
        return self.archivo

    def cargar(self):
        self.buscar_excel()
        self.workbook = load_workbook(self.archivo, data_only=True)
        self.hojas = self.workbook.sheetnames
        self._reconstruir_aliases()
        self.escanear_excel_completo()
        return True

    def _normalizar_nombre(self, valor: str) -> str:
        texto = str(valor or "").strip().upper()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(c for c in texto if not unicodedata.combining(c))
        texto = texto.replace("-", "_").replace(" ", "_")
        texto = re.sub(r"[^A-Z0-9_]+", "", texto)
        texto = re.sub(r"_+", "_", texto).strip("_")
        return texto

    def _normalizar_texto(self, valor) -> str:
        if valor is None:
            return ""
        return str(valor).strip()

    def _normalizar_estado(self, valor: str) -> str:
        txt = self._normalizar_nombre(valor)
        mapping = {
            "COMPLETADO": "Completado",
            "CUMPLIDO": "Completado",
            "CERRADO": "Completado",
            "PENDIENTE": "Pendiente",
            "ABIERTO": "Pendiente",
            "EN_PROCESO": "En proceso",
            "EN_CURSO": "En proceso",
            "VENCIDO": "Vencido",
        }
        return mapping.get(txt, self._normalizar_texto(valor))

    def _normalizar_prioridad(self, valor: str) -> str:
        txt = self._normalizar_nombre(valor)
        mapping = {
            "CRITICA": "Critica",
            "CRITICO": "Critica",
            "ALTA": "Alta",
            "MEDIA": "Media",
            "BAJA": "Baja",
            "URGENTE": "Critica",
        }
        if txt in mapping:
            return mapping[txt]
        if txt.isdigit():
            n = int(txt)
            if n >= 90:
                return "Critica"
            if n >= 70:
                return "Alta"
            if n >= 40:
                return "Media"
            return "Baja"
        return self._normalizar_texto(valor)

    def _normalizar_numero(self, valor: str):
        txt = self._normalizar_texto(valor).replace(" ", "")
        if not txt:
            return ""
        if re.match(r"^-?\d+[\.,]\d+$", txt):
            txt2 = txt.replace(".", "").replace(",", ".") if txt.count(",") == 1 and txt.count(".") > 1 else txt.replace(",", ".")
            try:
                return float(txt2)
            except Exception:
                return txt
        if re.match(r"^-?\d+$", txt):
            try:
                return int(txt)
            except Exception:
                return txt
        return txt

    def _normalizar_hora(self, valor: str) -> str:
        txt = self._normalizar_texto(valor)
        if not txt:
            return ""
        if re.match(r"^\d{1,2}:\d{2}$", txt):
            h, m = txt.split(":")
            return f"{int(h):02d}:{int(m):02d}"
        if re.match(r"^\d{1,2}\.\d{2}$", txt):
            h, m = txt.split(".")
            return f"{int(h):02d}:{int(m):02d}"
        return txt

    def _normalizar_fecha_texto(self, valor: str) -> str:
        txt = self._normalizar_texto(valor)
        if not txt:
            return ""
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(txt, fmt).strftime("%Y-%m-%d")
            except Exception:
                pass
        return txt

    def _normalizar_valor(self, valor, key_hint: str = ""):
        if valor is None:
            return ""
        if isinstance(valor, datetime):
            if key_hint and ("hora" in key_hint or "ingreso" in key_hint or "salida" in key_hint or "colacion" in key_hint):
                return valor.strftime("%H:%M")
            return valor.strftime("%Y-%m-%d")
        if isinstance(valor, date):
            return valor.strftime("%Y-%m-%d")
        if isinstance(valor, time):
            return valor.strftime("%H:%M")
        if isinstance(valor, (int, float)):
            return valor

        txt = self._normalizar_texto(valor)
        if not txt:
            return ""

        key = self._normalizar_nombre(key_hint).lower()
        if any(k in key for k in ["estado"]):
            return self._normalizar_estado(txt)
        if any(k in key for k in ["prioridad"]):
            return self._normalizar_prioridad(txt)
        if any(k in key for k in ["fecha", "venc", "plazo"]):
            return self._normalizar_fecha_texto(txt)
        if any(k in key for k in ["hora", "ingreso", "salida", "colacion", "llegada"]):
            return self._normalizar_hora(txt)

        n = self._normalizar_numero(txt)
        if isinstance(n, (int, float)):
            return n
        return txt

    def _reconstruir_aliases(self):
        self.alias_hojas = {}
        if not self.hojas:
            return

        norm_to_real = {self._normalizar_nombre(h): h for h in self.hojas}
        for canon, variantes in self.ALIASES_CRITICOS.items():
            elegido = None
            for variante in variantes:
                v_norm = self._normalizar_nombre(variante)
                if v_norm in norm_to_real:
                    elegido = norm_to_real[v_norm]
                    break
                for n_real, real in norm_to_real.items():
                    if v_norm and (v_norm in n_real or n_real in v_norm):
                        elegido = real
                        break
                if elegido:
                    break
            if elegido:
                self.alias_hojas[canon] = elegido

    def resolver_hoja(self, nombre: str) -> str | None:
        if not nombre:
            return None
        if nombre in self.hojas:
            return nombre

        norm = self._normalizar_nombre(nombre)
        for real in self.hojas:
            if self._normalizar_nombre(real) == norm:
                return real

        for canon, real in self.alias_hojas.items():
            if canon == nombre or self._normalizar_nombre(canon) == norm:
                return real

        for real in self.hojas:
            n_real = self._normalizar_nombre(real)
            if norm and (norm in n_real or n_real in norm):
                return real
        return None

    def existe_hoja(self, nombre: str) -> bool:
        return self.resolver_hoja(nombre) is not None

    def existe_alias(self, alias_canonico: str) -> bool:
        return alias_canonico in self.alias_hojas and bool(self.alias_hojas.get(alias_canonico))

    def obtener_hoja(self, nombre: str):
        real = self.resolver_hoja(nombre)
        if real and self.workbook is not None:
            return self.workbook[real]
        return None

    def leer_celda(self, hoja: str, celda: str):
        ws = self.obtener_hoja(hoja)
        if ws is None:
            return None
        return ws[celda].value

    def _analizar_headers(self, ws, max_rows: int = 20):
        candidatos = []
        max_col = ws.max_column or 1

        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_col, values_only=True), start=1):
            texts = []
            for c in row:
                t = self._normalizar_texto(c)
                if t:
                    texts.append(t)

            if not texts:
                continue

            norm_texts = [self._normalizar_nombre(t).lower() for t in texts]
            unique_ratio = len(set(norm_texts)) / max(1, len(norm_texts))
            keyword_score = sum(1 for t in norm_texts if any(k in t for k in self.KEYWORDS_HEADERS))
            score = len(texts) + keyword_score * 2 + int(unique_ratio * 10)

            if len(texts) >= 2:
                candidatos.append({
                    "fila": idx,
                    "headers": texts,
                    "score": score,
                    "keyword_score": keyword_score,
                })

        candidatos.sort(key=lambda x: x["score"], reverse=True)
        mejor = candidatos[0] if candidatos else None
        return mejor, candidatos[:3]

    def detectar_encabezados(self, hoja):
        ws = self.obtener_hoja(hoja)
        if ws is None:
            return None, []
        return self._analizar_headers(ws, max_rows=20)

    def clasificar_hoja(self, nombre_hoja):
        ws = self.obtener_hoja(nombre_hoja)
        if ws is None:
            return "Otros"

        nombre_norm = self._normalizar_nombre(ws.title)
        header, _ = self.detectar_encabezados(ws.title)
        headers = header.get("headers", []) if header else []
        headers_norm = [self._normalizar_nombre(h) for h in headers]
        name_only = nombre_norm

        def has_any(tokens):
            return any(self._normalizar_nombre(t) in name_only for t in tokens)

        # 1) Prioridad absoluta por nombre exacto de hoja
        if nombre_norm in self.EXCEL_DOMAIN_OVERRIDES:
            return self.EXCEL_DOMAIN_OVERRIDES[nombre_norm]

        # 2) Palabras del nombre de hoja (antes de prefijos y encabezados)
        if has_any(["VISION_BOARD"]):
            return "Vision Board"
        if has_any(["DESARROLLO_PERSONAL"]):
            return "Desarrollo Personal"
        if has_any(["SERPAT", "TURNO", "TURNOS"]):
            return "SERPAT"
        if has_any(["PENDIENTE", "PENDIENTES"]):
            return "Pendientes"
        if has_any(["ANCLA"]):
            return "Ancla"
        if has_any(["KPI"]):
            return "KPI"
        if has_any(["ALERTA"]):
            return "Alertas"

        # Reglas obligatorias de dominio por nombre
        # Finanzas (con prioridad sobre inversiones para patrimonio_neto)
        if has_any(["PATRIMONIO_NETO", "CONCILIACION", "INGRESOS", "GASTOS", "DEUDAS", "CONTROL_PAGOS", "BALANCE", "FINANZAS"]):
            return "Finanzas"

        # Inversiones (excepto patrimonio_neto ya capturado arriba)
        if has_any(["INVERSIONES", "RENTABILIDAD", "POSICION", "PORTFOLIO", "TQQQ", "SOXX", "PATRIMONIAL"]):
            return "Inversiones"

        # Salud
        if has_any(["SALUD", "HEALTH", "MEDICO", "PRESION", "MEDICAMENTOS", "RECETAS", "EXAMENES", "CHECKUP", "DENTAL", "FAMILIAR"]):
            return "Salud"

        # Empresas
        if has_any(["EMPRESA", "EMPRESAS", "MGC", "LNR", "LNRE", "CAPTAPROPIA", "CRM", "MERCADO", "COMPETENCIA", "CAMPANAS", "REDES", "WEB", "HOLDING"]):
            return "Empresas"

        # Universidad (excepto auditorias explicitas)
        if has_any(["UNIVERSIDAD", "RAMOS", "RAMO", "EVALUACIONES", "EVALUACION", "NOTAS", "CALENDARIO_ACADEMICO", "BLOQUES_ESTUDIO", "ERRORES_APRENDIZAJE", "REPASO", "T2"]):
            if not has_any(["AUDITORIA", "AUDIT"]):
                return "Universidad"

        if has_any(["HISTORIAL", "RESPALDO"]):
            return "Historial"

        # 3) Prefijos numéricos conocidos y ordenados
        for patron, dominio in self.PREFIX_DOMAIN_RULES:
            if re.match(patron, nombre_norm):
                return dominio

        # 4) Keywords como último recurso (si el nombre no fue concluyente)
        corpus = " ".join([nombre_norm] + headers_norm)
        score = {d: 0 for d in self.DOMINIOS_FUNCIONALES}
        for dominio, keywords in self.DOMINIO_KEYWORDS.items():
            for kw in keywords:
                kw_norm = self._normalizar_nombre(kw)
                if kw_norm and kw_norm in corpus:
                    score[dominio] += 1

        # Distinguir Inversiones dentro de Finanzas cuando no hay prefijo determinante
        if score.get("Inversiones", 0) >= 1 and score.get("Finanzas", 0) >= 1:
            return "Inversiones"

        best = max(score.items(), key=lambda x: x[1])
        return best[0] if best[1] > 0 else "Otros"

    def escanear_hoja(self, nombre):
        ws = self.obtener_hoja(nombre)
        if ws is None:
            return {
                "nombre": nombre,
                "error": "Hoja no encontrada",
            }

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        filas_con_datos = 0
        cols_con_datos = set()

        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True), start=1):
            tiene_dato = False
            for c_idx, value in enumerate(row, start=1):
                if self._normalizar_texto(value):
                    tiene_dato = True
                    cols_con_datos.add(c_idx)
            if tiene_dato:
                filas_con_datos += 1

        header_principal, header_candidatos = self._analizar_headers(ws, max_rows=20)
        tabla = self.detectar_tabla(nombre)
        dominio = self.clasificar_hoja(ws.title)
        estado_lectura = "OK"
        if filas_con_datos == 0:
            estado_lectura = "Vacia"
        elif not header_principal:
            estado_lectura = "Advertencia"

        return {
            "nombre": ws.title,
            "dimensiones": f"{max_row}x{max_col}",
            "cantidad_filas": max_row,
            "cantidad_columnas": max_col,
            "filas_con_datos": filas_con_datos,
            "columnas_con_datos": len(cols_con_datos),
            "posibles_encabezados": header_candidatos,
            "encabezado_detectado": header_principal,
            "tablas_detectadas": [tabla] if tabla else [],
            "categoria_funcional": dominio,
            "estado_lectura": estado_lectura,
            "vacia": filas_con_datos == 0,
        }

    def detectar_tabla(self, nombre):
        ws = self.obtener_hoja(nombre)
        if ws is None:
            return None

        header, _ = self._analizar_headers(ws)
        if not header:
            return None

        start_row = header["fila"]
        headers = [self._normalizar_texto(h) for h in header["headers"]]
        max_row = ws.max_row or start_row
        rows_data = 0
        end_row = start_row

        for r in range(start_row + 1, max_row + 1):
            row_vals = [self._normalizar_texto(v) for v in ws[r]]
            if any(row_vals):
                rows_data += 1
                end_row = r
            else:
                if rows_data > 0:
                    break

        return {
            "hoja": ws.title,
            "fila_encabezado": start_row,
            "encabezados": headers,
            "fila_inicio_datos": start_row + 1,
            "fila_fin_datos": end_row,
            "filas_utiles": rows_data,
        }

    def leer_tabla_inteligente(self, nombre):
        ws = self.obtener_hoja(nombre)
        if ws is None:
            return []

        tabla = self.detectar_tabla(nombre)
        if not tabla:
            return []

        head_row = tabla["fila_encabezado"]
        start = tabla["fila_inicio_datos"]
        end = tabla["fila_fin_datos"]
        headers_raw = [self._normalizar_texto(v) for v in ws[head_row]]

        headers = []
        for i, h in enumerate(headers_raw, start=1):
            headers.append(h if h else f"col_{i}")

        data = []
        for r in range(start, end + 1):
            row = ws[r]
            item = {}
            has_any = False
            for i, cell in enumerate(row, start=1):
                key = headers[i - 1] if i - 1 < len(headers) else f"col_{i}"
                value = self._normalizar_valor(cell.value, key_hint=key)
                if value != "":
                    has_any = True
                item[key] = value
            if has_any:
                item["_hoja"] = ws.title
                data.append(item)
        return data

    def buscar_hojas_por_categoria(self, categoria):
        categoria_norm = self._normalizar_nombre(categoria).lower()
        resultados = []

        # Permite buscar por dominio funcional completo
        for dominio, hojas in self.dominios.items():
            if self._normalizar_nombre(dominio).lower() == categoria_norm:
                return list(hojas)

        canonicos = self.CATEGORIAS.get(categoria_norm, [])
        for canon in canonicos:
            real = self.alias_hojas.get(canon)
            if real and real not in resultados:
                resultados.append(real)

        if resultados:
            return resultados

        for h in self.hojas:
            n = self._normalizar_nombre(h).lower()
            if categoria_norm and categoria_norm in n:
                resultados.append(h)
        return resultados

    def leer_ultimos_registros(self, nombre, cantidad=5):
        tabla = self.leer_tabla_inteligente(nombre)
        if not tabla:
            return []
        cantidad = max(1, int(cantidad or 1))
        return tabla[-cantidad:]

    def buscar_valores(self, nombre, palabras_clave):
        tabla = self.leer_tabla_inteligente(nombre)
        if not tabla:
            return []

        claves = [self._normalizar_nombre(p).lower() for p in (palabras_clave or []) if str(p).strip()]
        if not claves:
            return []

        encontrados = []
        for row in tabla:
            texto = " ".join(str(v) for v in row.values() if v is not None)
            texto_norm = self._normalizar_nombre(texto).lower()
            if any(k in texto_norm for k in claves):
                encontrados.append(row)
        return encontrados

    def escanear_excel_completo(self):
        self.scan = {}
        self.catalogo = []
        self.dominios = {k: [] for k in self.DOMINIOS_FUNCIONALES}
        self.hojas_vacias = []
        self.hojas_problema = []
        self.advertencias = []

        for h in self.hojas:
            try:
                info = self.escanear_hoja(h)
                self.scan[h] = info
                self.catalogo.append(info)

                dominio = info.get("categoria_funcional", "Otros")
                if dominio not in self.dominios:
                    dominio = "Otros"
                self.dominios[dominio].append(h)

                if info.get("vacia"):
                    self.hojas_vacias.append(h)
                if not info.get("encabezado_detectado") and not info.get("vacia"):
                    self.hojas_problema.append(f"{h}: sin encabezado detectado")
                    self.advertencias.append(f"Encabezado no detectado en hoja {h}")
            except Exception as e:
                self.scan[h] = {"nombre": h, "error": str(e)}
                self.hojas_problema.append(f"{h}: {str(e)}")
                self.catalogo.append({
                    "nombre": h,
                    "cantidad_filas": 0,
                    "cantidad_columnas": 0,
                    "encabezado_detectado": None,
                    "filas_con_datos": 0,
                    "categoria_funcional": "Otros",
                    "estado_lectura": "Error",
                    "error": str(e),
                })
                self.dominios["Otros"].append(h)

        return self.catalogo

    def escanear_excel_inteligente(self):
        # Compatibilidad con integraciones previas
        return self.escanear_excel_completo()

    def obtener_dominio(self, nombre_dominio):
        dominio_norm = self._normalizar_nombre(nombre_dominio)
        for dominio, hojas in self.dominios.items():
            if self._normalizar_nombre(dominio) == dominio_norm:
                return list(hojas)
        return []

    def obtener_resumen_dominios(self):
        resumen = {}
        for dominio, hojas in self.dominios.items():
            filas = 0
            for h in hojas:
                filas += int(self.scan.get(h, {}).get("filas_con_datos", 0) or 0)
            resumen[dominio] = {
                "hojas": len(hojas),
                "filas_utiles": filas,
                "hojas_lista": list(hojas),
            }
        return resumen

    def buscar_en_excel(self, palabras_clave):
        claves = [self._normalizar_nombre(p).lower() for p in (palabras_clave or []) if str(p).strip()]
        if not claves:
            return []

        resultados = []
        for hoja in self.hojas:
            try:
                tabla = self.leer_tabla_inteligente(hoja)
            except Exception:
                tabla = []

            if not tabla:
                continue

            for row in tabla[:250]:
                texto = " ".join(str(v) for v in row.values() if v is not None)
                texto_norm = self._normalizar_nombre(texto).lower()
                if any(k in texto_norm for k in claves):
                    resultados.append({
                        "hoja": hoja,
                        "dominio": self.scan.get(hoja, {}).get("categoria_funcional", "Otros"),
                        "registro": row,
                    })
        return resultados

    def generar_reporte_excel(self, ruta="Salidas/Logs/EXCEL_INTELLIGENCE_REPORT.txt"):
        if not self.catalogo:
            self.escanear_excel_completo()

        path = Path(ruta)
        path.parent.mkdir(parents=True, exist_ok=True)

        resumen_dom = self.obtener_resumen_dominios()
        hojas_ok = [x for x in self.catalogo if x.get("estado_lectura") in {"OK", "Advertencia"}]

        lineas = []
        lineas.append("VIDA_REAL_ENGINE V5.4.4 FINAL - EXCEL INTELLIGENCE LAYER")
        lineas.append("=" * 78)
        lineas.append(f"Archivo: {self.archivo.name if self.archivo else 'N/D'}")
        lineas.append(f"Total de hojas: {len(self.hojas)}")
        lineas.append(f"Hojas leidas correctamente: {len(hojas_ok)}")
        lineas.append(f"Hojas vacias: {len(self.hojas_vacias)}")
        lineas.append(f"Hojas con problemas: {len(self.hojas_problema)}")
        lineas.append(f"Dominios detectados: {sum(1 for _, v in self.dominios.items() if v)}")
        lineas.append("")

        lineas.append("HOJAS POR DOMINIO")
        for dominio in self.DOMINIOS_FUNCIONALES:
            data = resumen_dom.get(dominio, {})
            lineas.append(f"- {dominio}: {data.get('hojas', 0)} hoja(s) | filas utiles: {data.get('filas_utiles', 0)}")

        lineas.append("")
        lineas.append("DOMINIOS CON MAYOR CANTIDAD DE INFORMACION")
        tops = sorted(resumen_dom.items(), key=lambda x: x[1].get("filas_utiles", 0), reverse=True)[:5]
        for dominio, data in tops:
            lineas.append(f"- {dominio}: {data.get('filas_utiles', 0)} filas utiles")

        lineas.append("")
        lineas.append("CATALOGO COMPLETO DE HOJAS")
        for info in self.catalogo:
            head = info.get("encabezado_detectado") or {}
            headers = head.get("headers", [])[:6] if isinstance(head, dict) else []
            lineas.append(f"- Hoja: {info.get('nombre', 'N/D')}")
            lineas.append(f"  filas: {info.get('cantidad_filas', 0)}")
            lineas.append(f"  columnas: {info.get('cantidad_columnas', 0)}")
            lineas.append(f"  encabezados detectados: {', '.join(headers) if headers else 'N/D'}")
            lineas.append(f"  filas utiles: {info.get('filas_con_datos', 0)}")
            lineas.append(f"  categoria funcional: {info.get('categoria_funcional', 'Otros')}")
            lineas.append(f"  estado de lectura: {info.get('estado_lectura', 'N/D')}")

        lineas.append("")
        lineas.append("HOJAS VACIAS")
        if self.hojas_vacias:
            for h in self.hojas_vacias:
                lineas.append(f"- {h}")
        else:
            lineas.append("- Ninguna")

        lineas.append("")
        lineas.append("HOJAS CON PROBLEMAS")
        if self.hojas_problema:
            for p in self.hojas_problema:
                lineas.append(f"- {p}")
        else:
            lineas.append("- Ninguna")

        lineas.append("")
        lineas.append("ADVERTENCIAS")
        if self.advertencias:
            for a in self.advertencias[:200]:
                lineas.append(f"- {a}")
        else:
            lineas.append("- Ninguna")

        path.write_text("\n".join(lineas), encoding="utf-8")
        return path

    def generar_reporte_scan(self, ruta="Salidas/Logs/excel_scan_report.txt"):
        # Compatibilidad; delega en reporte final de intelligence layer.
        return self.generar_reporte_excel(ruta)

    def resumen(self):
        return {
            "archivo": self.archivo.name if self.archivo else "",
            "total_hojas": len(self.hojas),
            "hojas": self.hojas,
            "hojas_alias": self.alias_hojas,
            "scan_total": len(self.scan),
            "hojas_vacias": len(self.hojas_vacias),
            "hojas_problema": len(self.hojas_problema),
            "dominios_detectados": sum(1 for _, hojas in self.dominios.items() if hojas),
            "dominios": {k: len(v) for k, v in self.dominios.items()},
        }
